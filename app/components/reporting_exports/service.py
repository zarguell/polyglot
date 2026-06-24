"""ReportService — CSV, XLSX (openpyxl), and PDF (weasyprint) backends."""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any

import structlog

logger = structlog.get_logger()


class ReportService:
    """Generate reports in multiple formats.

    Backends:
        csv  — built-in csv module
        xlsx — openpyxl
        pdf  — weasyprint (HTML-to-PDF)
    """

    def __init__(self, temp_dir: str | None = None) -> None:
        self._temp_dir = temp_dir or os.getenv("EXPORT_TEMP_DIR", "./exports")
        Path(self._temp_dir).mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        format: str,
        data: list[dict[str, Any]],
        filename: str,
        columns: list[str] | None = None,
        title: str | None = None,
    ) -> str:
        """Generate a report file and return its path.

        Args:
            format: One of 'csv', 'xlsx', 'pdf'.
            data: List of dicts representing rows.
            filename: Base filename (without extension).
            columns: Ordered list of column names (defaults to data keys).
            title: Report title (used for PDF).
        """
        columns = columns or (list(data[0].keys()) if data else [])
        filepath = os.path.join(self._temp_dir, f"{filename}.{format}")

        if format == "csv":
            self._generate_csv(filepath, data, columns)
        elif format == "xlsx":
            self._generate_xlsx(filepath, data, columns)
        elif format == "pdf":
            self._generate_pdf(filepath, data, columns, title or filename)
        else:
            raise ValueError(f"Unsupported format: {format}")

        logger.info("report_generated", format=format, filepath=filepath)
        return filepath

    def _generate_csv(self, filepath: str, data: list[dict[str, Any]], columns: list[str]) -> None:
        with open(filepath, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)

    def _generate_xlsx(self, filepath: str, data: list[dict[str, Any]], columns: list[str]) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Data rows
        for row_idx, row in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        wb.save(filepath)

    def _generate_pdf(
        self, filepath: str, data: list[dict[str, Any]], columns: list[str], title: str
    ) -> None:
        from weasyprint import HTML

        html = self._build_html_table(data, columns, title)
        HTML(string=html).write_pdf(filepath)

    @staticmethod
    def _build_html_table(data: list[dict[str, Any]], columns: list[str], title: str) -> str:
        """Build a simple HTML table for PDF rendering."""
        rows_html = ""
        for row in data:
            cells = "".join(f"<td>{row.get(col, '')}</td>" for col in columns)
            rows_html += f"<tr>{cells}</tr>"

        header_cells = "".join(f"<th>{col}</th>" for col in columns)

        return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>{title}</title>
<style>
    body {{ font-family: sans-serif; margin: 40px; }}
    h1 {{ color: #333; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
    th {{ background-color: #f2f2f2; }}
</style>
</head>
<body>
    <h1>{title}</h1>
    <table><thead><tr>{header_cells}</tr></thead><tbody>{rows_html}</tbody></table>
</body>
</html>"""
